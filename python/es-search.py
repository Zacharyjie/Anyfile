# -*- coding: utf-8 -*-

from elasticsearch import Elasticsearch
from openpyxl import Workbook
import re
import math
import json
import datetime
import base64
import os

class apixls:
    def __init__(self,index):
        es = Elasticsearch('IP:port',timeout=120)
        osds = os.getcwd()+'\\'+'bool.json'
        f = open(osds,'r',encoding='utf-8')
        boby = json.load(f)
        self.wb = Workbook()
        self.res = es.search(index=index, doc_type="doc", body=boby)
    
    def get_finchina(self):
        final_result =[]
        finchina = []
        for num in self.res['hits']['hits']:
            datetime = num['_source']['@timestamp'][:10]
            api_foo = num['_source']['foo']
            dizhi = num['_source']['client_ip']           
            yonghu = num['_source'].get('consumer')
            if yonghu is None:
                final_result.append("%s %s %s %s"%(datetime,yonghu,api_foo,dizhi))
            else:
                final_result.append("%s %s %s %s"%(datetime,yonghu['username'],api_foo,dizhi))
        my_set = set(final_result)
        for wum in my_set:
            finchina.append("%s %s"%(wum,final_result.count(wum)))
        return finchina
    
    def get_finchina_token(self):
        final_result =[]
        finchina = []
        for num in self.res['hits']['hits']:
            datetime = num['_source']['@timestamp'][:10]
            api_foo = num['_source']['foo']
            dizhi = num['_source']['client_ip']
            base = num['_source']['request']['querystring'].get('token')
            if base is None:
                final_result.append("%s %s %s %s"%(datetime,base,api_foo,dizhi))
            else:
                base_fit = base.split(".")[1]
                decode = str(base64.b64decode(base_fit+"=="))
                yonghu = decode.split(",")[1].split("\"")[3]
                final_result.append("%s %s %s %s"%(datetime,yonghu,api_foo,dizhi))
        #print (self.res)
        my_set = set(final_result)
        for wum in my_set:
            finchina.append("%s %s"%(wum,final_result.count(wum)))
        return finchina

    def write_xls(self,food):
        ws = self.wb.active
        ws.title = "企业api"
        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['D'].width = 27
        ws.column_dimensions['E'].width = 15.5
        titles = ['日期','用户','用户名','访问接口地','ip地址','次数']
        ws.append(titles)
        aum = 2
        for gum in food:
            final_num = gum.split(' ')
            if final_num[1] == 'None':
                ws.cell(row=aum, column=1, value=final_num[0])
                ws.cell(row=aum, column=4, value=final_num[2])
                ws.cell(row=aum, column=5, value=final_num[3])
                ws.cell(row=aum, column=6, value=final_num[4])
            else:
                ws.cell(row=aum, column=1, value=final_num[0])
                ws.cell(row=aum, column=2, value=final_num[1])
                ws.cell(row=aum, column=3, value=final_num[1])
                ws.cell(row=aum, column=4, value=final_num[2])
                ws.cell(row=aum, column=5, value=final_num[3])
                ws.cell(row=aum, column=6, value=final_num[4])
            aum = aum + 1
        self.wb.save('es-search.xlsx')

if __name__ == "__main__":
    #now = datetime.datetime.now()
    total = []
    today = datetime.date.today()
    #tweek_head = today - datetime.timedelta(days=today.weekday())
    #tweek_fri = today + datetime.timedelta(days=4-today.weekday())
    tweek_head = today + datetime.timedelta(days=today.weekday()-7)
    tweek_fri = today - datetime.timedelta(days=today.weekday())
    while tweek_head <= tweek_fri:
        num = tweek_head.strftime("%Y.%m.%d")
        index = 'logstash-api-%s'%(num)
        qiyexls = apixls(index)
        tweek_head += datetime.timedelta(days=1)
        get_value = qiyexls.get_finchina_token()
        for i in get_value:
            total.append(i)
        qiyexls.write_xls(total)