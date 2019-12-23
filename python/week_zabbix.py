# -*- coding: utf-8 -*-

import pymysql
import time
import datetime
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment, NamedStyle
from openpyxl import Workbook

class week_zabbix:
    def __init__(self,zabbix_sql,today,last_weekend,hard_weekend):
        self.db = pymysql.connect(host='ip',user='zabbix',password='zabbix',db='zabbix')
        self.sql = zabbix_sql
        self.wb = Workbook()
        self.today = time.mktime(time.strptime('{0} 00:00:00'.format(today), '%Y-%m-%d %H:%M:%S'))
        self.last_weekend = time.mktime(time.strptime('{0} 00:00:00'.format(last_weekend), '%Y-%m-%d %H:%M:%S'))
        self.hard_weekend = time.mktime(time.strptime('{0} 00:00:00'.format(hard_weekend), '%Y-%m-%d %H:%M:%S'))

    def zabbix_mysql(self):
        cursor = self.db.cursor()
        cursor.execute(self.sql.format(self.today,self.last_weekend,self.hard_weekend))
        results = cursor.fetchall()
        cursor.close()
        self.db.close()
        return results

    def write_xls(self,total,sheet_name,sheet_num):
        for num in range(sheet_num):
            ws = self.wb.create_sheet(sheet_name[num],index=num)
            main_fs = NamedStyle(name='main_{0}'.format(sheet_name[num]), font=Font(name='等线', size='11'), alignment=Alignment(horizontal='right'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
            top_fs  = NamedStyle(name='top_{0}'.format(sheet_name[num]), font=Font(name='等线', size='14', bold=True),alignment=Alignment(horizontal='center'))
            ws.column_dimensions['A'].width = 8.75
            ws.column_dimensions['B'].width = 15.5
            if sheet_name[num].find('流量') == -1:
                ws.column_dimensions['C'].width = 35.0
                ws.column_dimensions['D'].width = 13.2
                ws.column_dimensions['E'].width = 13.2
                ws.column_dimensions['F'].width = 13.2
                ws.column_dimensions['G'].width = 13.2
                ws.column_dimensions['H'].width = 13.2
            else:
                ws.column_dimensions['C'].width = 65.0
                ws.column_dimensions['D'].width = 14.2
                ws.column_dimensions['E'].width = 14.2
                ws.column_dimensions['F'].width = 14.2
                ws.column_dimensions['G'].width = 14.2
                ws.column_dimensions['H'].width = 14.2

            titles = ['Hostid','主机地址','监控项目名','本周最小值','本周最大值','本周平均值','上周平均值','增比']
            for i in range(len(titles)):
                ws.cell(row=1, column=(i+1), value=titles[i]).style = top_fs
            aum = 2
            for i in total[num]:
                ws.cell(row=aum, column=1, value=i[0]).style = main_fs
                ws.cell(row=aum, column=2, value=i[1]).style = main_fs
                ws.cell(row=aum, column=3, value=i[2]).style = main_fs
                ws.cell(row=aum, column=4, value=i[3]).style = main_fs
                ws.cell(row=aum, column=5, value=i[4]).style = main_fs
                ws.cell(row=aum, column=6, value=i[5]).style = main_fs
                ws.cell(row=aum, column=7, value=i[6]).style = main_fs
                ws.cell(row=aum, column=8, value=i[7]).style = main_fs
                aum = aum + 1
        self.wb.save('Zabbix本周统计.xlsx')

if __name__ == "__main__":
    total = []
    now = datetime.datetime.now()
    #今天日期
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    #上周日日期
    last_weekend = ( now - datetime.timedelta(days=now.weekday()+1)).strftime('%Y-%m-%d')
    #上上周日日期
    hard_weekend = ( now - datetime.timedelta(days=now.weekday()+8)).strftime('%Y-%m-%d')
    #提取监控的项目名
    sheet_name = ['Windows_Cpu','Windows_Mem','Windows_C盘','Windows_D盘','Windows_E盘','Windows入口流量','Windows出口流量','Linux_Cpu','Linux_Mem','Linux根目录','Linux入口流量','Linux出口流量']
    sql = [
        (
            "select a.hostid,a.name,b.name,"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='CPU使用率百分比' "
            "left join trends c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "left join trends d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid order by host"            
        ),#Windows_Cpu
        (
            "select a.hostid,a.name,b.name,"
            "(case when min(c.value_min)/(1024*1024*1024) >= 1.000 then concat(convert(min(c.value_min)/(1024*1024*1024),decimal(16,2)),' GB') else concat(convert(min(c.value_min)/(1024*1024),decimal(16,2)),' MB') end),"    
            "(case when max(c.value_max)/(1024*1024*1024) >= 1.000 then concat(convert(max(c.value_max)/(1024*1024*1024),decimal(16,2)),' GB') else concat(convert(max(c.value_max)/(1024*1024),decimal(16,2)),' MB') end),"
            "(case when avg(c.value_avg)/(1024*1024*1024) >= 1.000 then concat(convert(avg(c.value_avg)/(1024*1024*1024),decimal(16,2)),' GB') else concat(convert(avg(c.value_avg)/(1024*1024),decimal(16,2)),' MB') end),"
            "(case when avg(d.value_avg)/(1024*1024*1024) >= 1.000 then concat(convert(avg(d.value_avg)/(1024*1024*1024),decimal(16,2)),' GB') else concat(convert(avg(d.value_avg)/(1024*1024),decimal(16,2)),' MB') end),"
            "(case when abs(avg(c.value_avg)-avg(d.value_avg))/(1024*1024*1024) >= 1.000 then concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024*1024),decimal(16,2)),' GB') "
            "else concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024*1024),decimal(16,2)),' MB') end) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Free memory' "
            "LEFT JOIN trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "LEFT JOIN trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid ORDER BY host"            
        ),#Windows_mem
        (
            "select a.hostid,a.name,replace(b.name,'$1','C:'),"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Free disk space on $1 (percentage)' and b.key_ like '%C:%'"
            "LEFT JOIN trends c on c.itemid=b.itemid and c.clock between {1} and {0}"
            "LEFT JOIN trends d on d.itemid=b.itemid and d.clock between {2} and {1}"
            "group by c.itemid,d.itemid ORDER BY host"            
        ),#Windows_C盘
        (
            "select a.hostid,a.name,replace(b.name,'$1','D:'),"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Free disk space on $1 (percentage)' and b.key_ like '%D:%'"
            "LEFT JOIN trends c on c.itemid=b.itemid and c.clock between {1} and {0}"
            "LEFT JOIN trends d on d.itemid=b.itemid and d.clock between {2} and {1}"
            "group by c.itemid,d.itemid ORDER BY host"            
        ),#Windows_D盘
        (
            "select a.hostid,a.name,replace(b.name,'$1','E:'),"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Free disk space on $1 (percentage)' and b.key_ like '%E:%'"
            "LEFT JOIN trends c on c.itemid=b.itemid and c.clock between {1} and {0}"
            "LEFT JOIN trends d on d.itemid=b.itemid and d.clock between {2} and {1}"
            "group by c.itemid,d.itemid ORDER BY host"            
        ),#Windows_E盘
        (
            "select a.hostid,a.name,replace(b.name,'$1','Intel(R) PRO/1000 MT Network Connection'),"
            "(case when min(c.value_min)/(1024*1024) >= 1.000 then concat(convert(min(c.value_min)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(min(c.value_min)/(1024),decimal(16,2)),' Kbps') end),"    
            "(case when max(c.value_max)/(1024*1024) >= 1.000 then concat(convert(max(c.value_max)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(max(c.value_max)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(c.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(c.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(c.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(d.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(d.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(d.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when abs(avg(c.value_avg)-avg(d.value_avg))/(1024*1024) >= 1.000 then concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024),decimal(16,2)),' Mbps') "
            "else concat(convert((avg(c.value_avg)-avg(d.value_avg))/1024,decimal(16,2)),' Kbps') end) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Incoming network traffic on $1' and b.key_ like '%Intel(R) PRO/1000 MT Network Connection]' "
            "LEFT JOIN trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "LEFT JOIN trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid ORDER BY host"
        ),#Windows入口流量
        (            
            "select a.hostid,a.name,replace(b.name,'$1','Intel(R) PRO/1000 MT Network Connection'),"
            "(case when min(c.value_min)/(1024*1024) >= 1.000 then concat(convert(min(c.value_min)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(min(c.value_min)/(1024),decimal(16,2)),' Kbps') end),"    
            "(case when max(c.value_max)/(1024*1024) >= 1.000 then concat(convert(max(c.value_max)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(max(c.value_max)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(c.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(c.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(c.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(d.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(d.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(d.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when abs(avg(c.value_avg)-avg(d.value_avg))/(1024*1024) >= 1.000 then concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024),decimal(16,2)),' Mbps') "
            "else concat(convert((avg(c.value_avg)-avg(d.value_avg))/1024,decimal(16,2)),' Kbps') end) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Outgoing network traffic on $1' and b.key_ like '%Intel(R) PRO/1000 MT Network Connection]' "
            "LEFT JOIN trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "LEFT JOIN trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid ORDER BY host"
        ),#Windows出口流量
        (
            "select a.hostid,a.name,replace(b.name,'$2','idle'),"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='CPU $2 time' and b.key_ like '%idle%' "
            "left join trends c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "left join trends d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid order by host"            
        ),#Linux_Cpu
        (
            "select a.hostid,a.name,b.name,"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='memory percent' "
            "left join trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "left join trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid order by host"            
        ),#Linux_Mem
        (
            "select a.hostid,a.name,replace(b.name,'$1','/:'),"
            "convert(min(c.value_min),decimal(10,2)),"
            "convert(max(c.value_max),decimal(10,2)),"
            "convert(avg(c.value_avg),decimal(10,2)),"
            "convert(avg(d.value_avg),decimal(10,2)),"
            "convert((avg(c.value_avg)-avg(d.value_avg)),decimal(10,2)) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name='Free disk space on $1 (percentage)' and b.key_ like '%/,%'"
            "LEFT JOIN trends c on c.itemid=b.itemid and c.clock between {1} and {0}"
            "LEFT JOIN trends d on d.itemid=b.itemid and d.clock between {2} and {1}"
            "group by c.itemid,d.itemid ORDER BY host"            
        ),#Linux根目录
        (            
            "select a.hostid,a.name,replace(b.name,'$1',substring_index(substring_index(b.key_ ,'[',-1),']',1)),"
            "(case when min(c.value_min)/(1024*1024) >= 1.000 then concat(convert(min(c.value_min)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(min(c.value_min)/(1024),decimal(16,2)),' Kbps') end),"    
            "(case when max(c.value_max)/(1024*1024) >= 1.000 then concat(convert(max(c.value_max)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(max(c.value_max)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(c.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(c.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(c.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(d.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(d.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(d.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when abs(avg(c.value_avg)-avg(d.value_avg))/(1024*1024) >= 1.000 then concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024),decimal(16,2)),' Mbps') "
            "else concat(convert((avg(c.value_avg)-avg(d.value_avg))/1024,decimal(16,2)),' Kbps') end) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name = 'Incoming network traffic on $1' and (b.key_ like '%eth%' or b.key_ like '%ens%') "
            "LEFT JOIN trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "LEFT JOIN trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid ORDER BY host"
        ),#Linux入口流量
        (            
            "select a.hostid,a.name,replace(b.name,'$1',substring_index(substring_index(b.key_ ,'[',-1),']',1)),"
            "(case when min(c.value_min)/(1024*1024) >= 1.000 then concat(convert(min(c.value_min)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(min(c.value_min)/(1024),decimal(16,2)),' Kbps') end),"    
            "(case when max(c.value_max)/(1024*1024) >= 1.000 then concat(convert(max(c.value_max)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(max(c.value_max)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(c.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(c.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(c.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when avg(d.value_avg)/(1024*1024) >= 1.000 then concat(convert(avg(d.value_avg)/(1024*1024),decimal(16,2)),' Mbps') else concat(convert(avg(d.value_avg)/(1024),decimal(16,2)),' Kbps') end),"
            "(case when abs(avg(c.value_avg)-avg(d.value_avg))/(1024*1024) >= 1.000 then concat(convert((avg(c.value_avg)-avg(d.value_avg))/(1024*1024),decimal(16,2)),' Mbps') "
            "else concat(convert((avg(c.value_avg)-avg(d.value_avg))/1024,decimal(16,2)),' Kbps') end) "
            "from hosts a inner join items b on b.hostid=a.hostid and b.name = 'Outgoing network traffic on $1' and (b.key_ like '%eth%' or b.key_ like '%ens%') "
            "LEFT JOIN trends_uint c on c.itemid=b.itemid and c.clock between {1} and {0} "
            "LEFT JOIN trends_uint d on d.itemid=b.itemid and d.clock between {2} and {1} "
            "group by c.itemid,d.itemid ORDER BY host"
        )#Linux出口流量
    ]
    for key in sql:
        total.append(week_zabbix(key,today,last_weekend,hard_weekend).zabbix_mysql())
    week_zabbix(sql,today,last_weekend,hard_weekend).write_xls(total,sheet_name,len(sheet_name))